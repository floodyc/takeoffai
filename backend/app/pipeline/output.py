"""Report generation — XLSX and CSV output from reconciliation results."""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .reconcile import ReconciliationReport

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, size=11, name='Arial', color='FFFFFF')
DATA_FONT = Font(size=10, name='Arial')
TOTAL_FILL = PatternFill('solid', fgColor='D6E4F0')
TOTAL_FONT = Font(bold=True, size=11, name='Arial')


def write_xlsx(report: ReconciliationReport, path: Path, page_label: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f'LT Fixture Label Count{" — " + page_label if page_label else ""}'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')

    # Summary table
    row = 3
    for col, h in enumerate(['LT Type', 'Final Count', 'Pass 1 (Raw)', 'Delta', 'Notes'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
        c.alignment = Alignment(horizontal='center')

    for i, (lt, count) in enumerate(sorted(report.final_counts.items())):
        r = row + 1 + i
        raw = report.pass1_counts.get(lt, 0)
        delta = count - raw
        delta_str = f"+{delta}" if delta > 0 else str(delta) if delta != 0 else "—"

        for col, val in enumerate([lt, count, raw, delta_str, ""], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font, c.border = DATA_FONT, THIN
            if col in (2, 3):
                c.alignment = Alignment(horizontal='center')

    # Total row
    tr = row + len(report.final_counts) + 1
    for col, val in enumerate(['TOTAL', report.total, report.pass1_total,
                                report.total - report.pass1_total, ''], 1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font, c.fill, c.border = TOTAL_FONT, TOTAL_FILL, THIN
        if col in (2, 3, 4):
            c.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40

    # Reconciliation detail sheet
    ws2 = wb.create_sheet("Reconciliation")
    ws2['A1'] = 'Boundary Additions'
    ws2['A1'].font = Font(bold=True, size=12, name='Arial')
    ws2['A2'] = 'Labels found in boundary strips but missed by both adjacent cells'
    ws2['A2'].font = Font(size=9, name='Arial', italic=True)

    r = 4
    for col, h in enumerate(['Label', 'Source Strip', 'Position', 'Confidence'], 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN

    for i, det in enumerate(report.boundary_additions):
        for col, val in enumerate([det.label_type, det.source_key, det.position, det.confidence], 1):
            c = ws2.cell(row=r + 1 + i, column=col, value=val)
            c.font, c.border = DATA_FONT, THIN

    r2 = r + len(report.boundary_additions) + 3
    ws2.cell(row=r2, column=1, value='Boundary Removals (Deduplication)').font = Font(bold=True, size=12, name='Arial')

    r2 += 2
    for col, h in enumerate(['Label', 'Removed From', 'Reason'], 1):
        c = ws2.cell(row=r2, column=col, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN

    for i, (lt, cell_key, reason) in enumerate(report.boundary_removals):
        for col, val in enumerate([lt, cell_key, reason], 1):
            c = ws2.cell(row=r2 + 1 + i, column=col, value=val)
            c.font, c.border = DATA_FONT, THIN

    # Warnings sheet
    if report.warnings:
        ws3 = wb.create_sheet("Warnings")
        ws3['A1'] = 'Ambiguous Cases'
        ws3['A1'].font = Font(bold=True, size=12, name='Arial')
        for i, w in enumerate(report.warnings):
            ws3.cell(row=i + 3, column=1, value=w).font = DATA_FONT

    for s in [ws2]:
        s.column_dimensions['A'].width = 15
        s.column_dimensions['B'].width = 25
        s.column_dimensions['C'].width = 20
        s.column_dimensions['D'].width = 15

    wb.save(str(path))
    print(f"[output] Saved {path}")


def write_csv(report: ReconciliationReport, path: Path):
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['LT Type', 'Final Count', 'Pass 1 Raw', 'Delta'])
        for lt, count in sorted(report.final_counts.items()):
            raw = report.pass1_counts.get(lt, 0)
            w.writerow([lt, count, raw, count - raw])
        w.writerow(['TOTAL', report.total, report.pass1_total, report.total - report.pass1_total])
    print(f"[output] Saved {path}")


def print_summary(report: ReconciliationReport):
    print("\n" + "=" * 60)
    print("FINAL LT LABEL COUNT")
    print("=" * 60)
    print(f"{'Type':<10} {'Count':>6}  {'(Raw)':>6}  {'Δ':>4}")
    print("-" * 35)
    for lt, count in sorted(report.final_counts.items()):
        raw = report.pass1_counts.get(lt, 0)
        delta = count - raw
        d = f"+{delta}" if delta > 0 else str(delta) if delta != 0 else ""
        print(f"{lt:<10} {count:>6}  {raw:>6}  {d:>4}")
    print("-" * 35)
    print(f"{'TOTAL':<10} {report.total:>6}  {report.pass1_total:>6}  "
          f"{'+' if report.total > report.pass1_total else ''}{report.total - report.pass1_total}")

    if report.boundary_additions:
        print(f"\n+{len(report.boundary_additions)} labels recovered from boundary strips")
    if report.boundary_removals:
        print(f"-{len(report.boundary_removals)} duplicates removed at boundaries")
    if report.warnings:
        print(f"⚠ {len(report.warnings)} ambiguous cases (see report)")
    print()
